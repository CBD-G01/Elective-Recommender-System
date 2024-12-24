import os
import openpyxl

def IsScore(sheet, r, c):
    if sheet.cell(row=r + 1, column=c).value != 'NT' \
            and sheet.cell(row=r + 1, column=c).value != 'I' \
            and sheet.cell(row=r + 1, column=c).value != 'XX' \
            and sheet.cell(row=r + 1, column=c).value != None:
        return True
    else:
        return False

def scr_sub(sheet, u, i):
    return sheet.cell(row=u + 1, column=i).value

print("Collaborative Filtering User-Based Algorithm for Grade Prediction")
wb = openpyxl.load_workbook('C:/Users/SRINIVAS/Downloads/Capstone/Elective-Recommender-System(GAGAN CBD01)/Elective-Recommender-System(GAGAN CBD01)/code/data/student.xlsx')

sheetname = wb.sheetnames

print("Student's score matrix(roll no in first column)\n")
sheet = wb[sheetname[0]]
import os
import openpyxl
from collections import defaultdict

def is_score(sheet, row, col):
    cell_value = sheet.cell(row=row + 1, column=col).value
    return cell_value not in ['NT', 'I', 'XX', None]

def get_score(sheet, row, col):
    return sheet.cell(row=row + 1, column=col).value

def get_average_grades(sheet, student):
    scores = [get_score(sheet, student, col) for col in range(2, sheet.max_column + 1) if is_score(sheet, student, col)]
    return round(sum(scores) / len(scores), 3) if scores else 0

def calculate_similarity(u, v, sheet, avg_u):
    avg_v = get_average_grades(sheet, v)
    scores_u = [get_score(sheet, u, col) for col in range(2, sheet.max_column + 1) if is_score(sheet, u, col) and is_score(sheet, v, col)]
    scores_v = [get_score(sheet, v, col) for col in range(2, sheet.max_column + 1) if is_score(sheet, u, col) and is_score(sheet, v, col)]
    sum1 = sum((score_u - avg_u) * (score_v - avg_v) for score_u, score_v in zip(scores_u, scores_v))
    sum2 = sum((score_u - avg_u) ** 2 for score_u in scores_u)
    sum3 = sum((score_v - avg_v) ** 2 for score_v in scores_v)
    return round(sum1 / ((sum2 * sum3) ** 0.5), 3) if sum2 and sum3 else 0

def predict_grades(sheet, student):
    avg_u = get_average_grades(sheet, student)
    similar_users = [(v, calculate_similarity(student, v, sheet, avg_u)) for v in range(1, sheet.max_row) if v != student and calculate_similarity(student, v, sheet, avg_u) > 0]
    predicted_grades = {}
    for course in range(2, sheet.max_column + 1):
        scores = [get_score(sheet, v, course) for v, _ in similar_users if is_score(sheet, v, course)]
        avg_scores = [get_average_grades(sheet, v) for v, _ in similar_users if is_score(sheet, v, course)]
        similarities = [sim for _, sim in similar_users if is_score(sheet, _, course)]
        sum1 = sum(sim * (score - avg) for sim, score, avg in zip(similarities, scores, avg_scores))
        sum2 = sum(similarities)
        predicted_grade = round(sum1 / sum2, 3) + avg_u if sum2 else 0
        predicted_grades[sheet.cell(row=1, column=course).value] = predicted_grade
    return predicted_grades

def main():
    print("Collaborative Filtering User-Based Algorithm for Grade Prediction")
    wb = openpyxl.load_workbook('C:/Users/SRINIVAS/Downloads/Capstone/Elective-Recommender-System(GAGAN CBD01)/Elective-Recommender-System(GAGAN CBD01)/code/data/student.xlsx')
    sheetname = wb.sheetnames
    print("Student's score matrix(roll no in first column)\n")
    sheet = wb[sheetname[0]]
    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value is not None:
                if j == 1:
                    print(sheet.cell(row=i, column=j).value, end="  | " if int(sheet.cell(row=i, column=j).value / 10) == 0 else " |  ")
                else:
                    print(sheet.cell(row=i, column=j).value, end="  ")
                if sheet.cell(row=i, column=j).value not in ['I', 'XX', 'NT']:
                    print(end=" " if int(sheet.cell(row=i, column=j).value / 10) == 0 else "")
            else:
                print("-", end="   ")
        print()
    studid = int(input("Enter roll no. : "))
    courses_list = [sheet.cell(row=1, column=i).value for i in range(2, sheet.max_column + 1)]
    print("\n".join(courses_list))
    predicted_grades = predict_grades(sheet, studid)
    print("\nPredicted Grades for Student " + str(studid) + ":")
    for course_name, grade in predicted_grades.items():
        print(f"{course_name}: {grade}")

if __name__ == "__main__":
    main()
for i in range(2, sheet.max_row + 1, 1):
    for j in range(1, sheet.max_column + 1, 1):
        if sheet.cell(row=i, column=j).value != None:
            if j == 1:
                if int(sheet.cell(row=i, column=j).value / 10) == 0:
                    print(sheet.cell(row=i, column=j).value, end="  | ")
                else:
                    print(sheet.cell(row=i, column=j).value, end=" |  ")
            else:
                print(sheet.cell(row=i, column=j).value, end="  ")
            if sheet.cell(row=i, column=j).value != 'I' and sheet.cell(row=i, column=j).value != 'XX' and sheet.cell(row=i, column=j).value != 'NT':
                if int(sheet.cell(row=i, column=j).value / 10) == 0:
                    print(end=" ")
            if sheet.cell(row=i, column=j).value == 'I':
                print(end=" ")
        else:
            print("-", end="   ")
    print()

studid = int(input("Enter roll no. : "))
coursesList = []
for i in range(2, sheet.max_column + 1):
    if (i - 2) % 9 == 0:
        print()
    print(sheet.cell(row=1, column=i).value, end=", ")
    coursesList.append(sheet.cell(row=1, column=i).value)
print("\n")

# Calculate average grades for a student
def getAvgGrades(sheet, student):
    sum = 0
    cnt = 0
    for i in range(2, sheet.max_column):
        if IsScore(sheet, student, i):
            sum += int(scr_sub(sheet, student, i))
            cnt += 1
    if cnt != 0:
        return round(float(sum / cnt), 3)
    else:
        return 0

# Calculate similarity between two students
def similarity(u, v, sheet, avgU):
    sum1 = 0
    sum2 = 0
    sum3 = 0
    avgV = getAvgGrades(sheet, v)

    for i in range(2, sheet.max_column + 1):
        if IsScore(sheet, u, i) and IsScore(sheet, v, i):
            sum1 += (scr_sub(sheet, u, i) - avgU) * (scr_sub(sheet, v, i) - avgV)
            sum2 += (scr_sub(sheet, u, i) - avgU) * (scr_sub(sheet, u, i) - avgU)
            sum3 += (scr_sub(sheet, v, i) - avgV) * (scr_sub(sheet, v, i) - avgV)
    if sum2 != 0 and sum3 != 0:
        return round(float(sum1 / ((sum2 * sum3) ** (1 / 2))), 3)
    else:
        return 0

similar_user_val = []
simi_user = []

avgU = getAvgGrades(sheet, studid)
for i in range(1, sheet.max_row):
    if i != studid:
        sim = similarity(studid, i, sheet, avgU)
        similar_user_val.append([studid, i, sim])
        if sim > 0:
            simi_user.append([studid, i, sim])

# Predicting grades for all courses
predicted_grades = {}
for course in range(2, sheet.max_column + 1):
    sum1 = 0
    sum2 = 0
    for i in range(0, len(simi_user)):
        if IsScore(sheet, int(simi_user[i][1]), course):
            avg = getAvgGrades(sheet, simi_user[i][1])
            sum1 += simi_user[i][2] * (scr_sub(sheet, simi_user[i][1], course) - avg)
            sum2 += simi_user[i][2]
    if sum2 != 0:
        predicted_grade = round(float(sum1 / sum2), 3) + avgU
        predicted_grades[sheet.cell(row=1, column=course).value] = predicted_grade

# Displaying predicted grades for all courses
print("\nPredicted Grades for Student " + str(studid) + ":")
for course_name, grade in predicted_grades.items():
    print(f"{course_name}: {grade}")